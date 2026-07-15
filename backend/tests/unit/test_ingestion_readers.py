from datetime import date, datetime
from pathlib import Path

import pytest
from bi_system.ingestion.readers import (
    InvalidWorkbookError,
    UnsupportedCsvEncodingError,
    WorksheetNotFoundError,
    iter_csv_rows,
    iter_xlsx_rows,
    xlsx_sheet_names,
)
from openpyxl import Workbook


def test_csv_reader_streams_rows_with_chinese_headers(tmp_path: Path) -> None:
    csv_path = tmp_path / "cities.csv"
    csv_path.write_text("城市,数值\n北京,10\n上海,20\n", encoding="utf-8-sig")

    rows = iter_csv_rows(csv_path)

    assert next(rows) == ("城市", "数值")
    assert next(rows) == ("北京", "10")
    assert list(rows) == [("上海", "20")]


def test_csv_reader_supports_explicit_gb18030(tmp_path: Path) -> None:
    csv_path = tmp_path / "legacy.csv"
    csv_path.write_text("城市,数值\n广州,30\n", encoding="gb18030")

    assert list(iter_csv_rows(csv_path, encoding="gb18030")) == [
        ("城市", "数值"),
        ("广州", "30"),
    ]


def test_csv_reader_rejects_unknown_encoding(tmp_path: Path) -> None:
    csv_path = tmp_path / "cities.csv"
    csv_path.write_text("city,value\n", encoding="utf-8")

    with pytest.raises(UnsupportedCsvEncodingError, match="supported encodings"):
        next(iter_csv_rows(csv_path, encoding="latin-1"))


def test_xlsx_reader_uses_read_only_values_without_executing_formulas(tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "数据"
    worksheet.append(["城市", "日期", "公式"])
    worksheet.append(["北京", date(2026, 7, 15), "=1+1"])
    workbook.create_sheet("说明")
    workbook.save(workbook_path)
    workbook.close()

    assert xlsx_sheet_names(workbook_path) == ("数据", "说明")

    rows = list(iter_xlsx_rows(workbook_path, sheet_name="数据"))
    assert rows[0] == ("城市", "日期", "公式")
    assert rows[1][0] == "北京"
    assert isinstance(rows[1][1], datetime)
    assert rows[1][2] == "=1+1"


def test_xlsx_reader_rejects_unknown_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    workbook = Workbook()
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(WorksheetNotFoundError, match="does not exist"):
        next(iter_xlsx_rows(workbook_path, sheet_name="missing"))


def test_xlsx_reader_translates_corrupt_workbook_errors(tmp_path: Path) -> None:
    workbook_path = tmp_path / "broken.xlsx"
    workbook_path.write_bytes(b"not a zip archive")

    with pytest.raises(InvalidWorkbookError, match="invalid or unreadable"):
        xlsx_sheet_names(workbook_path)
