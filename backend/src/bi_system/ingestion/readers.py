import csv
from collections.abc import Generator, Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

SUPPORTED_CSV_ENCODINGS = frozenset({"utf-8", "utf-8-sig", "gb18030"})


class IngestionReaderError(ValueError):
    """Base error for input files that cannot be read safely."""


class UnsupportedCsvEncodingError(IngestionReaderError):
    pass


class InvalidWorkbookError(IngestionReaderError):
    pass


class WorksheetNotFoundError(IngestionReaderError):
    pass


def iter_csv_rows(path: Path, *, encoding: str = "utf-8-sig") -> Iterator[tuple[str, ...]]:
    normalized_encoding = encoding.lower()
    if normalized_encoding not in SUPPORTED_CSV_ENCODINGS:
        supported = ", ".join(sorted(SUPPORTED_CSV_ENCODINGS))
        msg = f"Unsupported CSV encoding {encoding!r}; supported encodings: {supported}"
        raise UnsupportedCsvEncodingError(msg)

    with path.open(encoding=normalized_encoding, newline="") as file_handle:
        for row in csv.reader(file_handle):
            yield tuple(row)


def xlsx_sheet_names(path: Path) -> tuple[str, ...]:
    with _open_xlsx_workbook(path) as workbook:
        return tuple(workbook.sheetnames)


def iter_xlsx_rows(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> Iterator[tuple[object, ...]]:
    with _open_xlsx_workbook(path) as workbook:
        selected_sheet = sheet_name or workbook.sheetnames[0]
        if selected_sheet not in workbook.sheetnames:
            msg = f"Worksheet {selected_sheet!r} does not exist"
            raise WorksheetNotFoundError(msg)

        worksheet = workbook[selected_sheet]
        for row in worksheet.iter_rows(values_only=True):
            yield tuple(row)


@contextmanager
def _open_xlsx_workbook(path: Path) -> Generator[Workbook]:
    file_handle: BinaryIO | None = None
    try:
        file_handle = path.open("rb")
        workbook = load_workbook(
            filename=file_handle,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        if file_handle is not None:
            file_handle.close()
        raise InvalidWorkbookError("The XLSX workbook is invalid or unreadable") from exc

    try:
        yield workbook
    finally:
        workbook.close()
        file_handle.close()
